# mcp_server.py  — replace the WHOLE file with this

import os
from pathlib import Path
from threading import Lock
from typing import List, Any, Dict, Optional

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill

app = FastAPI(title="Excel-MCP (single-file mode)")

# 🔒 **ONE fixed workbook only**
EXCEL_FILE = Path(__file__).parent / "sample_data.xlsx"
if not EXCEL_FILE.exists():
    Workbook().save(EXCEL_FILE)   # create empty file on first run

wb_lock = Lock()  # 🛡️ thread-safe writes


# ========= Schemas =========
class WriteDataRequest(BaseModel):
    data: List[List[Any]]
    sheet_name: str = "Sheet1"
    start_cell: str = "A1"


class FormatRequest(BaseModel):
    sheet_name: str
    cell_range: str            # e.g. "A1:C2"
    bold: Optional[bool] = False
    bg_color: Optional[str] = None  # hex, e.g. "#FFFF00"


# ========= Helpers =========
def save_workbook(wb):
    with wb_lock:
        wb.save(EXCEL_FILE)


# ========= Routes =========
@app.get("/")
async def root():
    return {
        "status": "ok",
        "message": "MCP server running — editing sample_data.xlsx only",
        "file": EXCEL_FILE.name,
    }


@app.post("/write-data")
async def write_to_excel(req: WriteDataRequest):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[req.sheet_name] if req.sheet_name in wb else wb.create_sheet(req.sheet_name)

        start_row = int("".join(filter(str.isdigit, req.start_cell)) or 1)
        start_col = (
            ord("".join(filter(str.isalpha, req.start_cell)).upper() or "A") - ord("A") + 1
        )

        for r_idx, row in enumerate(req.data, start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                ws.cell(row=r_idx, column=c_idx, value=value)

        save_workbook(wb)
        return {"status": "written", "file": EXCEL_FILE.name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/format-range")
async def format_excel_range(req: FormatRequest):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[req.sheet_name]

        for row in ws[req.cell_range]:
            for cell in row:
                if req.bold:
                    cell.font = Font(bold=True)
                if req.bg_color:
                    cell.fill = PatternFill(
                        start_color=req.bg_color.lstrip("#"),
                        end_color=req.bg_color.lstrip("#"),
                        fill_type="solid",
                    )
        save_workbook(wb)
        return {"status": "formatted", "file": EXCEL_FILE.name}
    except KeyError:
        raise HTTPException(status_code=404, detail="Sheet not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/download")
async def download():
    return FileResponse(
        EXCEL_FILE,
        filename=EXCEL_FILE.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

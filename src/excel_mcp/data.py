from pathlib import Path
from typing import Any
import logging

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .exceptions import DataError
from .cell_utils import parse_cell_range

logger = logging.getLogger(__name__)

def read_excel_range(
    filepath: Path | str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: str | None = None,
    preview_only: bool = False
) -> list[dict[str, Any]]:
    """Read data from Excel range with optional preview mode"""
    try:
        wb = load_workbook(filepath, read_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise DataError(f"Sheet '{sheet_name}' not found")
            
        ws = wb[sheet_name]

        # Parse start cell
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Get start coordinates
        try:
            start_coords = parse_cell_range(f"{start_cell}:{start_cell}")
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Determine end coordinates
        if end_cell:
            try:
                end_coords = parse_cell_range(f"{end_cell}:{end_cell}")
                if not end_coords or not all(coord is not None for coord in end_coords[:2]):
                    raise DataError(f"Invalid end cell reference: {end_cell}")
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Invalid end cell format: {str(e)}")
        else:
            # Dynamically expand range until all values are empty
            end_row, end_col = start_row, start_col
            while end_row <= ws.max_row and any(ws.cell(row=end_row, column=c).value is not None for c in range(start_col, ws.max_column + 1)):
                end_row += 1
            while end_col <= ws.max_column and any(ws.cell(row=r, column=end_col).value is not None for r in range(start_row, ws.max_row + 1)):
                end_col += 1
            end_row -= 1  # Adjust back to last non-empty row
            end_col -= 1  # Adjust back to last non-empty column

        # Validate range bounds
        if start_row > ws.max_row or start_col > ws.max_column:
            raise DataError(
                f"Start cell out of bounds. Sheet dimensions are "
                f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            )

        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            if any(v is not None for v in row_data):
                data.append(row_data)

        wb.close()
        return data
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to read Excel range: {e}")
        raise DataError(str(e))

def write_data(
    filepath: str,
    sheet_name: str | None,
    data: list[list] | None,
    start_cell: str = "A1",
) -> dict[str, str]:
    """Write data to Excel sheet with workbook handling
    
    Headers are handled intelligently based on context.
    """
    try:
        if not data:
            raise DataError("No data provided to write")
            
        wb = load_workbook(filepath)

        # If no sheet specified, use active sheet
        if not sheet_name:
            sheet_name = wb.active.title
        elif sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # Validate start cell
        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(coord is not None for coord in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        if len(data) > 0:
            _write_data_to_worksheet(ws, data, start_cell)

        wb.save(filepath)
        wb.close()

        return {"message": f"Data written to {sheet_name}", "active_sheet": sheet_name}
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        raise DataError(str(e))

def _looks_like_headers(row_dict):
    """Check if a data row appears to be headers (keys match values)."""
    return all(
        isinstance(value, str) and str(value).strip() == str(key).strip()
        for key, value in row_dict.items()
    )
    
def _check_for_headers_above(worksheet, start_row, start_col, headers):
    """Check if cells above start position contain headers."""
    if start_row <= 1:
        return False  # Nothing above row 1
        
    # Look for header-like content above
    for check_row in range(max(1, start_row - 5), start_row):
        # Count matches for this row
        header_count = 0
        cell_count = 0
        
        for i, header in enumerate(headers):
            if i >= 10:  # Limit check to first 10 columns for performance
                break
                
            cell = worksheet.cell(row=check_row, column=start_col + i)
            cell_count += 1
            
            # Check if cell is formatted like a header (bold)
            is_formatted = cell.font.bold if hasattr(cell.font, 'bold') else False
            
            # Check for any content that could be a header
            if cell.value is not None:
                # Case 1: Direct match with expected header
                if str(cell.value).strip().lower() == str(header).strip().lower():
                    header_count += 2  # Give higher weight to exact matches
                # Case 2: Any formatted cell with content
                elif is_formatted and cell.value:
                    header_count += 1
                # Case 3: Any cell with content in the first row we check
                elif check_row == max(1, start_row - 5):
                    header_count += 0.5
        
        # If we have a significant number of matching cells, consider it a header row
        if cell_count > 0 and header_count >= cell_count * 0.5:
            return True
            
    # No headers found above
    return False

def _determine_header_behavior(worksheet, start_row, start_col, data):
    """Determine if headers should be written based on context."""
    if not data:
        return False  # No data means no headers
        
    # Check if we're in the title area (rows 1-4)
    if start_row <= 4:
        return False  # Don't add headers in title area
    
    # If we already have data in the sheet, be cautious about adding headers
    if worksheet.max_row > 1:
        # Check if the target row already has content
        has_content = any(
            worksheet.cell(row=start_row, column=start_col + i).value is not None
            for i in range(min(5, len(data[0].keys())))
        )
        
        if has_content:
            return False  # Don't overwrite existing content with headers
        
        # Check if first row appears to be headers
        first_row_is_headers = _looks_like_headers(data[0])
        
        # Check extensively for headers above (up to 5 rows)
        has_headers_above = _check_for_headers_above(worksheet, start_row, start_col, list(data[0].keys()))
        
        # Be conservative - don't add headers if we detect headers above or the data has headers
        if has_headers_above or first_row_is_headers:
            return False
        
        # If we're appending data immediately after existing data, don't add headers
        if any(worksheet.cell(row=start_row-1, column=start_col + i).value is not None 
               for i in range(min(5, len(data[0].keys())))):
            return False
    
    # For completely new sheets or empty areas far from content, add headers
    return True

def _write_data_to_worksheet(
    worksheet: Worksheet, 
    data: list[list], 
    start_cell: str = "A1",
) -> None:
    """Write data to worksheet with intelligent header handling"""
    try:
        if not data:
            raise DataError("No data provided to write")

        try:
            start_coords = parse_cell_range(start_cell)
            if not start_coords or not all(x is not None for x in start_coords[:2]):
                raise DataError(f"Invalid start cell reference: {start_cell}")
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Invalid start cell format: {str(e)}")

        # Write data
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                worksheet.cell(row=start_row + i, column=start_col + j, value=val)
    except DataError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to write worksheet data: {e}")
        raise DataError(str(e))
